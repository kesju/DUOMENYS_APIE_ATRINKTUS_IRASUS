from __future__ import annotations

import json
import math
import re
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill

DATA_DIR = Path("/Users/kesju/DI/DUOMENYS/ATRINKTI_ANOTUOTI_IRASAI")
OUTPUT_PATH = Path(
    "/Users/kesju/DI/CODEX_PROJECTS/DUOMENYS_APIE_ATRINKTUS_IRASUS/"
    "results/selected_records_metadata_v4.xlsx"
)
OUTPUT_COLUMNS = [
    "file_name",
    "recordingId",
    "userId",
    "comment",
    "mark",
    "FULLY_ANNOTATED_PROFESSIONALLY",
    "BAD_QUALITY",
    "NOISES_ANNOTATED",
    "Other Flags",
    "human",
    "ml",
]
PRIMARY_FLAGS = {
    "FULLY_ANNOTATED_PROFESSIONALLY",
    "BAD_QUALITY",
    "NOISES_ANNOTATED",
}
MARK_PATTERN = re.compile(r"(?<!\d)(1111|2222|3333|5555|9999)(?!\d)")


def extract_mark_from_comment(comment: str | None) -> str | None:
    """Extract the relevant mark code from a JSON comment if present."""
    if not comment:
        return None

    match = MARK_PATTERN.search(comment)
    return match.group(1) if match else None


def format_mapping(value: Any) -> str | None:
    """Format an annotation-count mapping for one Excel cell."""
    if not isinstance(value, dict):
        return None
    return json.dumps(value, ensure_ascii=False)


def extract_record_metadata(file_path: Path) -> tuple[dict[str, object], list[str]]:
    """Extract the requested metadata from one JSON record."""
    with file_path.open("r", encoding="utf-8") as file_handle:
        payload = json.load(file_handle)

    if not isinstance(payload, dict):
        return (
            {column: None for column in OUTPUT_COLUMNS} | {"file_name": file_path.stem},
            ["recordingId", "userId", "comment", "flags", "rpeakAnnotationCounts"],
        )

    missing_fields: list[str] = []
    recording_id = payload.get("recordingId")
    user_id = payload.get("userId")
    comment = payload.get("comment")

    flags_value = payload.get("flags")
    if isinstance(flags_value, list):
        flags = [str(flag) for flag in flags_value]
    else:
        flags = []
        missing_fields.append("flags")
    flag_set = set(flags)
    other_flags = [flag for flag in flags if flag not in PRIMARY_FLAGS]

    annotation_counts = payload.get("rpeakAnnotationCounts")
    if isinstance(annotation_counts, dict):
        human = format_mapping(annotation_counts.get("human"))
        ml = format_mapping(annotation_counts.get("ml"))
        if human is None:
            missing_fields.append("rpeakAnnotationCounts.human")
        if ml is None:
            missing_fields.append("rpeakAnnotationCounts.ml")
    else:
        human = None
        ml = None
        missing_fields.append("rpeakAnnotationCounts")

    if recording_id is None:
        missing_fields.append("recordingId")
    if user_id is None:
        missing_fields.append("userId")
    if comment is None:
        missing_fields.append("comment")

    row: dict[str, object] = {
        "file_name": file_path.stem,
        "recordingId": recording_id,
        "userId": user_id,
        "comment": comment,
        "mark": extract_mark_from_comment(comment),
        "FULLY_ANNOTATED_PROFESSIONALLY": (
            "FULLY_ANNOTATED_PROFESSIONALLY" in flag_set
        ),
        "BAD_QUALITY": "BAD_QUALITY" in flag_set,
        "NOISES_ANNOTATED": "NOISES_ANNOTATED" in flag_set,
        "Other Flags": (
            json.dumps(other_flags, ensure_ascii=False) if other_flags else None
        ),
        "human": human,
        "ml": ml,
    }
    return row, missing_fields


def estimate_comment_height(comment: object) -> float:
    """Estimate a readable height for a wrapped comment cell."""
    line_count = sum(
        max(1, math.ceil(len(line) / 75))
        for line in str(comment or "").splitlines() or [""]
    )
    return min(90, 15 * line_count)


def save_metadata(rows: list[dict[str, object]]) -> None:
    """Write metadata to a readable, filterable Excel worksheet."""
    dataframe = pd.DataFrame(rows, columns=OUTPUT_COLUMNS)

    with pd.ExcelWriter(OUTPUT_PATH, engine="openpyxl") as writer:
        dataframe.to_excel(writer, sheet_name="metadata", index=False)
        worksheet = writer.sheets["metadata"]
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
        worksheet.row_dimensions[1].height = 32

        column_widths = {
            "A": 22,
            "B": 28,
            "C": 28,
            "D": 80,
            "E": 12,
            "F": 34,
            "G": 18,
            "H": 26,
            "I": 38,
            "J": 28,
            "K": 28,
        }
        for column, width in column_widths.items():
            worksheet.column_dimensions[column].width = width

        for row_number in range(2, worksheet.max_row + 1):
            worksheet.cell(row_number, 1).number_format = "@"
            for column_number in range(1, len(OUTPUT_COLUMNS) + 1):
                worksheet.cell(row_number, column_number).alignment = Alignment(
                    horizontal=("center" if column_number in (6, 7, 8) else "left"),
                    vertical="top",
                    wrap_text=column_number in (4, 9, 10, 11),
                )

            worksheet.row_dimensions[row_number].height = estimate_comment_height(
                worksheet.cell(row_number, 4).value
            )


def main() -> None:
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)

    rows: list[dict[str, object]] = []
    missing_files: list[tuple[str, list[str]]] = []

    for file_path in sorted(DATA_DIR.rglob("*.json")):
        if file_path.name == "manifest.json":
            continue

        row, missing_fields = extract_record_metadata(file_path)
        rows.append(row)
        if missing_fields:
            missing_files.append((file_path.name, missing_fields))

    save_metadata(rows)

    print(f"Saved {len(rows)} rows to {OUTPUT_PATH}")
    print("Columns:", OUTPUT_COLUMNS)
    if missing_files:
        print("Files missing metadata:")
        for file_name, missing_fields in missing_files:
            print(f"- {file_name}: {', '.join(missing_fields)}")
    else:
        print("All JSON files contain the requested metadata.")


if __name__ == "__main__":
    main()
