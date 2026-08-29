from __future__ import annotations

import json
import math
import re
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill

DATA_DIR = Path("/Users/kesju/DI/DUOMENYS/ATRINKTI_ANOTUOTI_IRASAI")
OUTPUT_PATH = Path(
    "/Users/kesju/DI/CODEX_PROJECTS/DUOMENYS_APIE_ATRINKTUS_IRASUS/results/selected_records_metadata_v2.xlsx"
)
OUTPUT_COLUMNS = ["file_name", "recordingId", "userId", "comment", "mark"]
MARK_PATTERN = re.compile(r"(?<!\d)(1111|2222|3333|5555|9999)(?!\d)")


def extract_mark_from_comment(comment: str | None) -> str | None:
    """Extract the relevant mark code from a JSON comment if present."""
    if not comment:
        return None

    match = MARK_PATTERN.search(comment)
    if match is None:
        return None
    return match.group(1)


def extract_record_metadata(
    file_path: Path,
) -> tuple[str, str | None, str | None, str | None, list[str]]:
    """Return the requested metadata and any fields missing from one JSON record."""
    with file_path.open("r", encoding="utf-8") as fh:
        payload = json.load(fh)

    if not isinstance(payload, dict):
        return file_path.name, None, None, None, ["recordingId", "userId", "comment"]

    missing_fields: list[str] = []
    recording_id = payload.get("recordingId")
    user_id = payload.get("userId")
    comment = payload.get("comment")
    mark = extract_mark_from_comment(comment)

    if recording_id is None:
        missing_fields.append("recordingId")
    if user_id is None:
        missing_fields.append("userId")
    if comment is None:
        missing_fields.append("comment")

    return file_path.name, recording_id, user_id, comment, mark, missing_fields


def save_metadata(rows: list[dict[str, str | None]]) -> None:
    """Write metadata to a readable, filterable Excel worksheet."""
    df = pd.DataFrame(rows, columns=OUTPUT_COLUMNS)

    with pd.ExcelWriter(OUTPUT_PATH, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="metadata", index=False)
        worksheet = writer.sheets["metadata"]
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        column_widths = {"A": 22, "B": 28, "C": 28, "D": 100, "E": 12}
        for column, width in column_widths.items():
            worksheet.column_dimensions[column].width = width

        for row_number in range(2, worksheet.max_row + 1):
            for column_number in range(1, 6):
                worksheet.cell(row_number, column_number).alignment = Alignment(
                    vertical="top", wrap_text=column_number in (4, 5)
                )

            comment = str(worksheet.cell(row_number, 4).value or "")
            line_count = sum(
                max(1, math.ceil(len(line) / 95)) for line in comment.splitlines() or [""]
            )
            worksheet.row_dimensions[row_number].height = min(90, 15 * line_count)


def main() -> None:
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)

    rows: list[dict[str, str | None]] = []
    missing_files: list[tuple[str, list[str]]] = []

    for file_path in sorted(DATA_DIR.rglob("*.json")):
        if file_path.name == "manifest.json":
            continue

        file_name, recording_id, user_id, comment, mark, missing_fields = extract_record_metadata(
            file_path
        )
        rows.append(
            {
                "file_name": file_name,
                "recordingId": recording_id,
                "userId": user_id,
                "comment": comment,
                "mark": mark,
            }
        )

        if missing_fields:
            missing_files.append((file_name, missing_fields))

    save_metadata(rows)

    print(f"Saved {len(rows)} rows to {OUTPUT_PATH}")

    if missing_files:
        print("Files missing metadata:")
        for file_name, missing_fields in missing_files:
            print(f"- {file_name}: {', '.join(missing_fields)}")
    else:
        print("All JSON files contain recordingId, userId and comment.")


if __name__ == "__main__":
    main()
