"""Append metadata for newly added ECG JSON records to the existing workbook.

The existing rows in ``selected_records_metadata_with_notes.xlsx`` are used as
the immutable baseline. Only JSON basenames that are absent from that workbook
are appended, and the result is saved under a new filename.
"""

from __future__ import annotations

from copy import copy
from pathlib import Path

from openpyxl import load_workbook

if __package__:
    from . import create_selected_records_metadata_excel_v4_2 as base
else:
    import create_selected_records_metadata_excel_v4_2 as base


PROJECT_DIR = Path(
    "/Users/kesju/DI/CODEX_PROJECTS/DUOMENYS_APIE_ATRINKTUS_IRASUS"
)
DATA_DIR = Path("/Users/kesju/DI/DUOMENYS/ATRINKTI_ANOTUOTI_IRASAI")
SOURCE_XLSX = PROJECT_DIR / "results" / "selected_records_metadata_with_notes.xlsx"
OUTPUT_PATH = PROJECT_DIR / "results" / "selected_records_metadata_updated_1.xlsx"
SHEET_NAME = "metadata"
OUTPUT_COLUMNS = [*base.OUTPUT_COLUMNS, "notes"]


def normalize_basename(value: object) -> str:
    """Return the workbook/JSON basename without a trailing .json suffix."""
    basename = str(value).strip()
    if basename.casefold().endswith(".json"):
        basename = basename[:-5]
    return basename


def relevant_json_files() -> list[Path]:
    """Return sorted JSON inputs, excluding every manifest*.* filename."""
    return sorted(
        (
            file_path
            for file_path in DATA_DIR.rglob("*")
            if file_path.is_file()
            and file_path.suffix.casefold() == ".json"
            and not base.is_manifest(file_path)
        ),
        key=lambda file_path: str(file_path).casefold(),
    )


def validate_headers(worksheet: object) -> None:
    """Fail rather than append values under an unexpected column layout."""
    actual_columns = [
        worksheet.cell(row=1, column=column_number).value
        for column_number in range(1, worksheet.max_column + 1)
    ]
    if actual_columns != OUTPUT_COLUMNS:
        raise ValueError(
            "Unexpected source-workbook columns. "
            f"Expected {OUTPUT_COLUMNS}, found {actual_columns}."
        )


def existing_basenames(worksheet: object) -> set[str]:
    """Read and validate the basenames already present in the workbook."""
    basenames: list[str] = []
    for row_number in range(2, worksheet.max_row + 1):
        value = worksheet.cell(row=row_number, column=1).value
        if value is None or not str(value).strip():
            raise ValueError(f"Missing basename in existing Excel row {row_number}.")
        basenames.append(normalize_basename(value))

    duplicates = sorted(
        basename for basename in set(basenames) if basenames.count(basename) > 1
    )
    if duplicates:
        raise ValueError(f"Duplicate basenames in source workbook: {duplicates}")
    return set(basenames)


def copy_template_format(worksheet: object, source_row: int, target_row: int) -> None:
    """Copy established cell formatting, but never source-row values/comments."""
    for column_number in range(1, len(OUTPUT_COLUMNS) + 1):
        source_cell = worksheet.cell(row=source_row, column=column_number)
        target_cell = worksheet.cell(row=target_row, column=column_number)
        if source_cell.has_style:
            target_cell._style = copy(source_cell._style)


def main() -> None:
    if not SOURCE_XLSX.exists():
        raise FileNotFoundError(f"Missing source workbook: {SOURCE_XLSX}")
    if not DATA_DIR.is_dir():
        raise FileNotFoundError(f"Missing JSON input directory: {DATA_DIR}")

    workbook = load_workbook(SOURCE_XLSX)
    if SHEET_NAME not in workbook.sheetnames:
        raise ValueError(
            f"The source workbook has no {SHEET_NAME!r} sheet: {SOURCE_XLSX}"
        )
    worksheet = workbook[SHEET_NAME]
    validate_headers(worksheet)

    original_last_row = worksheet.max_row
    workbook_basenames = existing_basenames(worksheet)
    json_files = relevant_json_files()

    json_files_by_basename: dict[str, Path] = {}
    duplicate_json_basenames: list[str] = []
    for file_path in json_files:
        basename = normalize_basename(file_path.name)
        if basename in json_files_by_basename:
            duplicate_json_basenames.append(basename)
        else:
            json_files_by_basename[basename] = file_path
    if duplicate_json_basenames:
        raise ValueError(
            "Duplicate JSON basenames in the input directory: "
            f"{sorted(set(duplicate_json_basenames))}"
        )

    new_files = [
        file_path
        for basename, file_path in json_files_by_basename.items()
        if basename not in workbook_basenames
    ]
    new_files.sort(key=lambda file_path: str(file_path).casefold())

    missing_files: list[tuple[str, list[str]]] = []
    for file_path in new_files:
        row, missing_fields = base.extract_record_metadata(file_path)
        row["notes"] = None

        target_row = worksheet.max_row + 1
        copy_template_format(worksheet, original_last_row, target_row)
        for column_number, column_name in enumerate(OUTPUT_COLUMNS, start=1):
            worksheet.cell(row=target_row, column=column_number).value = row[column_name]

        worksheet.row_dimensions[target_row].height = base.estimate_comment_height(
            row["comment"]
        )
        if missing_fields:
            missing_files.append((file_path.name, missing_fields))

    if worksheet.auto_filter.ref:
        worksheet.auto_filter.ref = f"A1:S{worksheet.max_row}"

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT_PATH)

    print(f"Read {original_last_row - 1} existing rows from {SOURCE_XLSX}")
    print(f"Appended {len(new_files)} new rows")
    print(f"Saved {worksheet.max_row - 1} rows to {OUTPUT_PATH}")
    if missing_files:
        print("New JSON files missing metadata:")
        for file_name, missing_fields in missing_files:
            print(f"- {file_name}: {', '.join(missing_fields)}")
    else:
        print("All new JSON files contain the requested metadata.")


if __name__ == "__main__":
    main()
