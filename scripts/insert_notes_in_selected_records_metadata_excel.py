from __future__ import annotations

import math
import re
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

PROJECT_DIR = Path(
    "/Users/kesju/DI/CODEX_PROJECTS/DUOMENYS_APIE_ATRINKTUS_IRASUS"
)
DATA_DIR = Path("/Users/kesju/DI/DUOMENYS/ATRINKTI_ANOTUOTI_IRASAI")
SOURCE_XLSX = PROJECT_DIR / "results" / "selected_records_metadata_v4_2.xlsx"
NOTES_PATH = PROJECT_DIR / "results" / "selected_records_notes.txt"
OUTPUT_PATH = (
    PROJECT_DIR / "results" / "selected_records_metadata_with_notes.xlsx"
)

SOURCE_COLUMNS = [
    "basename",
    "recordingId",
    "userId",
    "comment",
    "tag",
    "rhythm",
    "h_nz_frac%",
    "N",
    "hS",
    "hV",
    "hU",
    "mlS",
    "mlV",
    "mlU",
    "FULLY_ANNOTATED_PROFESSIONALLY",
    "BAD_QUALITY",
    "NOISES_ANNOTATED",
    "Other Flags",
]
OUTPUT_COLUMNS = [*SOURCE_COLUMNS, "notes"]
FILE_NAME_RE = re.compile(r"^\s*(\d{7}\.\d{3})(?:\.json)?\s*$")
OWNER_RECORD_RE = re.compile(
    r"(?<![0-9A-Fa-f])[0-9A-Fa-f]{24}-\d+(?:-r)?(?![\w-])"
)
EMAIL_RE = re.compile(
    r"(?<![\w.+-])[\w.+-]+@[\w.-]+\.[A-Za-z]{2,}(?![\w.-])"
)


def normalize_basename(value: object) -> str:
    """Normalize an Excel or JSON name to the XXXXXXX.XXX identifier."""
    basename = str(value).strip()
    if basename.casefold().endswith(".json"):
        basename = basename[:-5]
    return basename


def is_manifest(path: Path) -> bool:
    """Return True for names matching manifest*.* (case-insensitive)."""
    return path.name.casefold().startswith("manifest") and "." in path.name


def parse_notes_file(path: Path) -> dict[str, str]:
    """Map each XXXXXXX.XXX heading to the text until the next heading."""
    notes_by_basename: dict[str, str] = {}
    current_basename: str | None = None
    current_lines: list[str] = []

    def save_current_block() -> None:
        if current_basename is None:
            return
        if current_basename in notes_by_basename:
            raise ValueError(
                f"Duplicate notes heading {current_basename!r} in {path}."
            )
        notes_by_basename[current_basename] = "\n".join(current_lines).strip()

    for source_line in path.read_text(encoding="utf-8").splitlines():
        match = FILE_NAME_RE.fullmatch(source_line)
        if match:
            save_current_block()
            current_basename = match.group(1)
            current_lines = []
        elif current_basename is not None:
            current_lines.append(source_line.rstrip())

    save_current_block()
    if not notes_by_basename:
        raise ValueError(f"No XXXXXXX.XXX headings found in {path}.")
    return notes_by_basename


def relevant_json_basenames() -> set[str]:
    """Return basenames of all JSON record files except manifest*.* files."""
    return {
        path.stem
        for path in DATA_DIR.rglob("*.json")
        if not is_manifest(path)
    }


def normalized_text(value: object) -> str:
    """Normalize whitespace for reliable text comparisons."""
    return " ".join(str(value or "").split())


def flexible_whitespace_pattern(text: str) -> re.Pattern[str] | None:
    """Build a literal text pattern that tolerates whitespace differences."""
    tokens = text.split()
    if not tokens:
        return None
    return re.compile(r"\s+".join(re.escape(token) for token in tokens))


def remove_comment_text(notes: str, comment: object) -> str:
    """Remove workbook-comment text while retaining additional note content."""
    comment_text = str(comment or "").strip()
    if not comment_text or not notes:
        return notes.strip()

    cleaned = notes

    # Prefer removing the complete comment, including multiline comments.
    full_pattern = flexible_whitespace_pattern(comment_text)
    if full_pattern is not None:
        cleaned = full_pattern.sub("", cleaned)

    # A notes section may contain only selected lines from a multiline comment.
    # Remove matching substantial lines, while short markers such as "Ž." are
    # removed only when they occupy a line by themselves.
    for comment_line in comment_text.splitlines():
        normalized_comment_line = normalized_text(comment_line)
        if not normalized_comment_line:
            continue

        if len(normalized_comment_line) >= 8:
            line_pattern = flexible_whitespace_pattern(comment_line)
            if line_pattern is not None:
                cleaned = line_pattern.sub("", cleaned)
        else:
            cleaned = "\n".join(
                line
                for line in cleaned.splitlines()
                if normalized_text(line) != normalized_comment_line
            )

    cleaned = re.sub(r"[ \t]+\n", "\n", cleaned)
    cleaned = re.sub(r"\n{3,}", "\n\n", cleaned)
    return cleaned.strip(" \t\n,;-–—")


def remove_owner_identifiers(notes: str) -> str:
    """Remove owner record identifiers and email addresses from notes."""
    cleaned_lines: list[str] = []
    for source_line in notes.splitlines():
        line = OWNER_RECORD_RE.sub("", source_line)
        line = EMAIL_RE.sub("", line)
        line = re.sub(r"[ \t]{2,}", " ", line).strip()
        line = re.sub(r"^[,;:–—-]+\s*", "", line)
        line = re.sub(r"\s*[,;:–—-]+$", "", line)
        cleaned_lines.append(line)

    cleaned = "\n".join(cleaned_lines)
    cleaned = re.sub(r"\n{3,}", "\n\n", cleaned)
    return cleaned.strip()


def compose_notes_cell(basename: str, note_body: str, comment: object) -> str:
    """Put the basename first, followed by notes not duplicated in comment."""
    cleaned_body = remove_comment_text(note_body, comment)
    cleaned_body = remove_owner_identifiers(cleaned_body)
    return f"{basename}\n{cleaned_body}" if cleaned_body else basename


def estimate_row_height(text: str, characters_per_line: int = 95) -> float:
    """Estimate a readable row height for wrapped multiline notes."""
    line_count = sum(
        max(1, math.ceil(len(line) / characters_per_line))
        for line in text.splitlines() or [""]
    )
    return min(150, 15 * line_count)


def workbook_basenames(worksheet: object, basename_column: int) -> list[str]:
    """Read and validate record basenames from the metadata worksheet."""
    basenames = [
        normalize_basename(worksheet.cell(row=row_number, column=basename_column).value)
        for row_number in range(2, worksheet.max_row + 1)
    ]
    duplicates = sorted(
        basename for basename in set(basenames) if basenames.count(basename) > 1
    )
    if duplicates:
        raise ValueError(f"Duplicate basenames in source workbook: {duplicates}")
    return basenames


def main() -> None:
    if not SOURCE_XLSX.exists():
        raise FileNotFoundError(f"Missing source workbook: {SOURCE_XLSX}")
    if not NOTES_PATH.exists():
        raise FileNotFoundError(f"Missing notes file: {NOTES_PATH}")

    notes_by_basename = parse_notes_file(NOTES_PATH)
    json_basenames = relevant_json_basenames()

    workbook = load_workbook(SOURCE_XLSX)
    if "metadata" not in workbook.sheetnames:
        raise ValueError(f"Source workbook has no 'metadata' sheet: {SOURCE_XLSX}")
    worksheet = workbook["metadata"]

    source_headers = [
        worksheet.cell(row=1, column=column_number).value
        for column_number in range(1, len(SOURCE_COLUMNS) + 1)
    ]
    if source_headers != SOURCE_COLUMNS:
        raise ValueError(
            f"Unexpected source columns. Expected {SOURCE_COLUMNS}, "
            f"found {source_headers}."
        )

    basename_column = SOURCE_COLUMNS.index("basename") + 1
    comment_column = SOURCE_COLUMNS.index("comment") + 1
    basenames = workbook_basenames(worksheet, basename_column)
    workbook_basename_set = set(basenames)

    if workbook_basename_set != json_basenames:
        raise ValueError(
            "The source workbook and relevant JSON files do not match. "
            f"Missing from workbook: {sorted(json_basenames - workbook_basename_set)}; "
            f"missing JSON files: {sorted(workbook_basename_set - json_basenames)}."
        )

    missing_notes = sorted(workbook_basename_set - set(notes_by_basename))
    if missing_notes:
        raise ValueError(f"No notes sections found for records: {missing_notes}")

    notes_column = len(OUTPUT_COLUMNS)
    notes_header = worksheet.cell(row=1, column=notes_column, value="notes")
    source_header = worksheet.cell(row=1, column=notes_column - 1)
    notes_header._style = copy(source_header._style)
    notes_header.alignment = copy(source_header.alignment)

    for row_number, basename in enumerate(basenames, start=2):
        comment = worksheet.cell(row=row_number, column=comment_column).value
        note_text = compose_notes_cell(
            basename,
            notes_by_basename[basename],
            comment,
        )
        notes_cell = worksheet.cell(
            row=row_number,
            column=notes_column,
            value=note_text,
        )
        notes_cell.alignment = Alignment(vertical="top", wrap_text=True)
        worksheet.row_dimensions[row_number].height = max(
            worksheet.row_dimensions[row_number].height or 15,
            estimate_row_height(note_text),
        )

    worksheet.column_dimensions[get_column_letter(notes_column)].width = 100
    worksheet.auto_filter.ref = (
        f"A1:{get_column_letter(notes_column)}{worksheet.max_row}"
    )

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT_PATH)

    unused_notes = sorted(set(notes_by_basename) - workbook_basename_set)
    print(f"Saved {len(basenames)} rows to {OUTPUT_PATH}")
    print("Columns:", OUTPUT_COLUMNS)
    print("Unused notes sections:", len(unused_notes))


if __name__ == "__main__":
    main()
