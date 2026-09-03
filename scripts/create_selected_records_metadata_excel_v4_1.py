from __future__ import annotations

import json
import math
import re
from pathlib import Path
from typing import Any

import pandas as pd
from ecg_io.extract_annotated_noise_metadata import extract_annotated_noise
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

DATA_DIR = Path("/Users/kesju/DI/DUOMENYS/ATRINKTI_ANOTUOTI_IRASAI")
OUTPUT_PATH = Path(
    "/Users/kesju/DI/CODEX_PROJECTS/DUOMENYS_APIE_ATRINKTUS_IRASUS/"
    "results/selected_records_metadata_v4_1.xlsx"
)

# All selected ECG records contain 127,999 int32 samples (511,996 bytes).
# Supplying this value lets
# the ECG pipeline calculate a percentage even though this directory contains
# only metadata JSON files and no same-stem signal files.
RECORD_SAMPLES = 127_999

OUTPUT_COLUMNS = [
    "basename",
    "recordingId",
    "userId",
    "comment",
    "tag",
    "rhythm",
    "h_nz_frac%",
    "FULLY_ANNOTATED_PROFESSIONALLY",
    "BAD_QUALITY",
    "NOISES_ANNOTATED",
    "Other Flags",
    "N",
    "hS",
    "hV",
    "hU",
    "mlS",
    "mlV",
    "mlU",
]
PRIMARY_FLAGS = {
    "FULLY_ANNOTATED_PROFESSIONALLY",
    "BAD_QUALITY",
    "NOISES_ANNOTATED",
}
TAG_PATTERN = re.compile(r"(?<!\d)(1111|2222|3333|5555|9999)(?!\d)")
AFIB_NOTE = "prieširdžių virpėjimas"


def extract_tag_from_comment(comment: str | None) -> str | None:
    """Extract the relevant tag code from a JSON comment if present."""
    if not comment:
        return None

    match = TAG_PATTERN.search(comment)
    return match.group(1) if match else None


def extract_rhythm_from_comment(comment: str | None) -> str | None:
    """Return AFIB when the Lithuanian AF description occurs in a comment."""
    if comment and AFIB_NOTE.casefold() in comment.casefold():
        return "AFIB"
    return None


def annotation_count(mapping: Any, annotation: str) -> int:
    """Return one integer annotation count, defaulting missing values to zero."""
    if not isinstance(mapping, dict):
        return 0

    value = mapping.get(annotation, 0)
    if isinstance(value, bool):
        return 0
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return 0


def calculate_h_nz_fraction(file_path: Path) -> float | None:
    """Calculate human-annotated noise as percent of record samples."""
    noise_metadata = extract_annotated_noise(file_path, samples=RECORD_SAMPLES)
    fraction = noise_metadata["noise_fraction_percent"]
    return round(float(fraction), 1) if fraction is not None else None


def extract_record_metadata(file_path: Path) -> tuple[dict[str, object], list[str]]:
    """Extract the requested metadata from one JSON record."""
    with file_path.open("r", encoding="utf-8") as file_handle:
        payload = json.load(file_handle)

    if not isinstance(payload, dict):
        return (
            {column: None for column in OUTPUT_COLUMNS}
            | {
                "basename": file_path.stem,
                "N": 0,
                "hS": 0,
                "hV": 0,
                "hU": 0,
                "mlS": 0,
                "mlV": 0,
                "mlU": 0,
            },
            ["recordingId", "userId", "comment", "flags", "rpeakAnnotationCounts"],
        )

    missing_fields: list[str] = []
    recording_id = payload.get("recordingId")
    user_id = payload.get("userId")
    comment_value = payload.get("comment")
    comment = comment_value if isinstance(comment_value, str) else None

    flags_value = payload.get("flags")
    if isinstance(flags_value, list):
        flags = [str(flag) for flag in flags_value]
    else:
        flags = []
        missing_fields.append("flags")
    flag_set = set(flags)
    other_flags = [flag for flag in flags if flag not in PRIMARY_FLAGS]

    annotation_counts = payload.get("rpeakAnnotationCounts")
    if isinstance(annotation_counts, dict):
        human_counts = annotation_counts.get("human")
        ml_counts = annotation_counts.get("ml")
        if not isinstance(human_counts, dict):
            human_counts = {}
            missing_fields.append("rpeakAnnotationCounts.human")
        if not isinstance(ml_counts, dict):
            ml_counts = {}
            missing_fields.append("rpeakAnnotationCounts.ml")
    else:
        human_counts = {}
        ml_counts = {}
        missing_fields.append("rpeakAnnotationCounts")

    if recording_id is None:
        missing_fields.append("recordingId")
    if user_id is None:
        missing_fields.append("userId")
    if comment_value is None:
        missing_fields.append("comment")

    row: dict[str, object] = {
        "basename": file_path.stem,
        "recordingId": recording_id,
        "userId": user_id,
        "comment": comment_value,
        "tag": extract_tag_from_comment(comment),
        "rhythm": extract_rhythm_from_comment(comment),
        "h_nz_frac%": calculate_h_nz_fraction(file_path),
        "FULLY_ANNOTATED_PROFESSIONALLY": (
            "FULLY_ANNOTATED_PROFESSIONALLY" in flag_set
        ),
        "BAD_QUALITY": "BAD_QUALITY" in flag_set,
        "NOISES_ANNOTATED": "NOISES_ANNOTATED" in flag_set,
        "Other Flags": (
            json.dumps(other_flags, ensure_ascii=False) if other_flags else None
        ),
        "N": annotation_count(human_counts, "N"),
        "hS": annotation_count(human_counts, "S"),
        "hV": annotation_count(human_counts, "V"),
        "hU": annotation_count(human_counts, "U"),
        "mlS": annotation_count(ml_counts, "S"),
        "mlV": annotation_count(ml_counts, "V"),
        "mlU": annotation_count(ml_counts, "U"),
    }
    return row, missing_fields


def estimate_comment_height(comment: object) -> float:
    """Estimate a readable height for a wrapped comment cell."""
    line_count = sum(
        max(1, math.ceil(len(line) / 75))
        for line in str(comment or "").splitlines() or [""]
    )
    return min(90, 15 * line_count)


def save_metadata(rows: list[dict[str, object]]) -> None:
    """Write metadata to a readable, filterable Excel worksheet."""
    dataframe = pd.DataFrame(rows, columns=OUTPUT_COLUMNS)

    with pd.ExcelWriter(OUTPUT_PATH, engine="openpyxl") as writer:
        dataframe.to_excel(writer, sheet_name="metadata", index=False)
        worksheet = writer.sheets["metadata"]
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
        worksheet.row_dimensions[1].height = 32

        column_widths = {
            "basename": 22,
            "recordingId": 28,
            "userId": 28,
            "comment": 80,
            "tag": 12,
            "rhythm": 12,
            "h_nz_frac%": 14,
            "FULLY_ANNOTATED_PROFESSIONALLY": 34,
            "BAD_QUALITY": 18,
            "NOISES_ANNOTATED": 26,
            "Other Flags": 38,
            "N": 10,
            "hS": 10,
            "hV": 10,
            "hU": 10,
            "mlS": 10,
            "mlV": 10,
            "mlU": 10,
        }
        for column_number, column_name in enumerate(OUTPUT_COLUMNS, start=1):
            column_letter = get_column_letter(column_number)
            worksheet.column_dimensions[column_letter].width = column_widths[column_name]

        centered_columns = {
            "tag",
            "rhythm",
            "h_nz_frac%",
            "FULLY_ANNOTATED_PROFESSIONALLY",
            "BAD_QUALITY",
            "NOISES_ANNOTATED",
            "N",
            "hS",
            "hV",
            "hU",
            "mlS",
            "mlV",
            "mlU",
        }
        wrapped_columns = {"comment", "Other Flags"}
        for row_number in range(2, worksheet.max_row + 1):
            worksheet.cell(row_number, 1).number_format = "@"
            for column_number, column_name in enumerate(OUTPUT_COLUMNS, start=1):
                cell = worksheet.cell(row_number, column_number)
                cell.alignment = Alignment(
                    horizontal=(
                        "center" if column_name in centered_columns else "left"
                    ),
                    vertical="top",
                    wrap_text=column_name in wrapped_columns,
                )
                if column_name == "h_nz_frac%":
                    cell.number_format = "0.0"

            worksheet.row_dimensions[row_number].height = estimate_comment_height(
                worksheet.cell(row_number, OUTPUT_COLUMNS.index("comment") + 1).value
            )


def is_manifest(file_path: Path) -> bool:
    """Return True for names matching manifest*.* (case-insensitive)."""
    return file_path.name.casefold().startswith("manifest") and "." in file_path.name


def main() -> None:
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)

    rows: list[dict[str, object]] = []
    missing_files: list[tuple[str, list[str]]] = []

    for file_path in sorted(DATA_DIR.rglob("*.json")):
        if is_manifest(file_path):
            continue

        row, missing_fields = extract_record_metadata(file_path)
        rows.append(row)
        if missing_fields:
            missing_files.append((file_path.name, missing_fields))

    save_metadata(rows)

    print(f"Saved {len(rows)} rows to {OUTPUT_PATH}")
    print("Columns:", OUTPUT_COLUMNS)
    if missing_files:
        print("Files missing metadata:")
        for file_name, missing_fields in missing_files:
            print(f"- {file_name}: {', '.join(missing_fields)}")
    else:
        print("All JSON files contain the requested metadata.")


if __name__ == "__main__":
    main()
