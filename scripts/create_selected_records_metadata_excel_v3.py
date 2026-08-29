from __future__ import annotations

import math
import re
from copy import copy
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment

SOURCE_XLSX = Path(
    "/Users/kesju/DI/CODEX_PROJECTS/DUOMENYS_APIE_ATRINKTUS_IRASUS/results/selected_records_metadata_v2.xlsx"
)
NOTES_PATH = Path(
    "/Users/kesju/DI/CODEX_PROJECTS/DUOMENYS_APIE_ATRINKTUS_IRASUS/results/selected_records_notes.txt"
)
OUTPUT_PATH = Path(
    "/Users/kesju/DI/CODEX_PROJECTS/DUOMENYS_APIE_ATRINKTUS_IRASUS/results/selected_records_metadata_v3.xlsx"
)
OUTPUT_COLUMNS = ["file_name", "recordingId", "userId", "comment", "mark", "notes"]
NOTE_FILE_NAME_RE = re.compile(r"^\s*(\d{7}\.\d{3})(?:\.json)?\s*$")


def normalize_file_name(file_name: str) -> str:
    value = str(file_name).strip()
    if value.endswith(".json"):
        value = value[:-5]
    return value


def parse_notes_file(path: Path) -> dict[str, str]:
    """Map each bare filename heading in the notes export to its note block."""
    text = path.read_text(encoding="utf-8")
    notes_by_file: dict[str, str] = {}
    current_file_name: str | None = None
    current_notes: list[str] = []

    def save_current_notes() -> None:
        if current_file_name is None:
            return

        first_content_line = 0
        last_content_line = len(current_notes)
        while first_content_line < last_content_line and not current_notes[
            first_content_line
        ].strip():
            first_content_line += 1
        while last_content_line > first_content_line and not current_notes[
            last_content_line - 1
        ].strip():
            last_content_line -= 1

        notes_by_file[normalize_file_name(current_file_name)] = "\n".join(
            line.rstrip() for line in current_notes[first_content_line:last_content_line]
        )

    for raw_line in text.splitlines():
        match = NOTE_FILE_NAME_RE.fullmatch(raw_line)
        if match:
            save_current_notes()
            current_file_name = match.group(1)
            current_notes = []
            continue

        if current_file_name is not None:
            current_notes.append(raw_line)

    save_current_notes()

    if not notes_by_file:
        raise ValueError(
            f"No filename headings in the form xxxxxxx.xxx were found in {path}."
        )

    return notes_by_file


def estimate_row_height(text: str, characters_per_line: int = 95) -> float:
    """Estimate a readable row height for wrapped multiline text."""
    line_count = sum(
        max(1, math.ceil(len(line) / characters_per_line))
        for line in str(text or "").splitlines() or [""]
    )
    return min(120, 15 * line_count)


def save_workbook(metadata_df: pd.DataFrame) -> None:
    """Add notes to a copy of v2 while preserving its established formatting."""
    workbook = load_workbook(SOURCE_XLSX)
    worksheet = workbook["metadata"]

    notes_column = 6
    header = worksheet.cell(row=1, column=notes_column, value="notes")
    source_header = worksheet.cell(row=1, column=5)
    header._style = copy(source_header._style)
    header.alignment = copy(source_header.alignment)

    for row_number, notes in enumerate(metadata_df["notes"], start=2):
        notes_cell = worksheet.cell(row=row_number, column=notes_column, value=notes)
        notes_cell.alignment = Alignment(vertical="top", wrap_text=True)
        worksheet.row_dimensions[row_number].height = max(
            worksheet.row_dimensions[row_number].height or 15,
            estimate_row_height(notes),
        )

    worksheet.column_dimensions["F"].width = 100
    worksheet.auto_filter.ref = f"A1:F{worksheet.max_row}"
    workbook.save(OUTPUT_PATH)


def main() -> None:
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)

    if not SOURCE_XLSX.exists():
        raise FileNotFoundError(f"Missing metadata source Excel: {SOURCE_XLSX}")
    if not NOTES_PATH.exists():
        raise FileNotFoundError(f"Missing notes source text: {NOTES_PATH}")

    metadata_df = pd.read_excel(SOURCE_XLSX)
    if "file_name" not in metadata_df.columns:
        raise ValueError(f"The file {SOURCE_XLSX} does not contain a file_name column.")

    notes_by_file = parse_notes_file(NOTES_PATH)
    metadata_df["notes"] = metadata_df["file_name"].map(
        lambda file_name: notes_by_file.get(normalize_file_name(file_name), "")
    )

    metadata_df = metadata_df[OUTPUT_COLUMNS]
    save_workbook(metadata_df)

    print(f"Saved {len(metadata_df)} rows to {OUTPUT_PATH}")
    print("Columns:", list(metadata_df.columns))
    missing_notes = metadata_df.loc[metadata_df["notes"] == "", "file_name"].tolist()
    print("Missing notes:", len(missing_notes))
    if missing_notes:
        print("Files without notes:")
        for file_name in missing_notes:
            print(f"- {file_name}")


if __name__ == "__main__":
    main()
