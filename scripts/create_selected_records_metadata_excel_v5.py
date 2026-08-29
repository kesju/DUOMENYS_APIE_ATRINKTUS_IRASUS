from __future__ import annotations

import math
import re
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment

PROJECT_DIR = Path(
    "/Users/kesju/DI/CODEX_PROJECTS/DUOMENYS_APIE_ATRINKTUS_IRASUS"
)
DATA_DIR = Path("/Users/kesju/DI/DUOMENYS/ATRINKTI_ANOTUOTI_IRASAI")
SOURCE_XLSX = PROJECT_DIR / "results" / "selected_records_metadata_v4.xlsx"
NOTES_PATH = PROJECT_DIR / "results" / "selected_records_notes.txt"
OUTPUT_PATH = PROJECT_DIR / "results" / "selected_records_metadata_v5.xlsx"

V4_COLUMNS = [
    "file_name",
    "recordingId",
    "userId",
    "comment",
    "mark",
    "FULLY_ANNOTATED_PROFESSIONALLY",
    "BAD_QUALITY",
    "NOISES_ANNOTATED",
    "Other Flags",
    "human",
    "ml",
]
OUTPUT_COLUMNS = [*V4_COLUMNS, "notes"]
FILE_NAME_RE = re.compile(r"^\s*(\d{7}\.\d{3})(?:\.json)?\s*$")


def normalize_file_name(value: object) -> str:
    """Normalize an Excel or JSON filename to the XXXXXXX.XXX identifier."""
    file_name = str(value).strip()
    if file_name.endswith(".json"):
        file_name = file_name[:-5]
    return file_name


def parse_notes_file(path: Path) -> dict[str, str]:
    """Map each filename heading in the notes file to all following note text."""
    notes_by_file: dict[str, str] = {}
    current_file_name: str | None = None
    current_lines: list[str] = []

    def save_current_block() -> None:
        if current_file_name is None:
            return

        first_line = 0
        last_line = len(current_lines)
        while first_line < last_line and not current_lines[first_line].strip():
            first_line += 1
        while last_line > first_line and not current_lines[last_line - 1].strip():
            last_line -= 1

        notes_by_file[current_file_name] = "\n".join(
            line.rstrip() for line in current_lines[first_line:last_line]
        )

    for source_line in path.read_text(encoding="utf-8").splitlines():
        match = FILE_NAME_RE.fullmatch(source_line)
        if match:
            save_current_block()
            current_file_name = match.group(1)
            current_lines = []
            continue

        if current_file_name is not None:
            current_lines.append(source_line)

    save_current_block()
    if not notes_by_file:
        raise ValueError(f"No XXXXXXX.XXX filename headings found in {path}.")
    return notes_by_file


def relevant_json_file_names() -> set[str]:
    """Return every relevant JSON filename as a normalized record identifier."""
    return {
        file_path.stem
        for file_path in DATA_DIR.rglob("*.json")
        if file_path.name != "manifest.json"
    }


def estimate_row_height(text: str, characters_per_line: int = 95) -> float:
    """Estimate a readable row height for wrapped multiline notes."""
    line_count = sum(
        max(1, math.ceil(len(line) / characters_per_line))
        for line in text.splitlines() or [""]
    )
    return min(120, 15 * line_count)


def main() -> None:
    if not SOURCE_XLSX.exists():
        raise FileNotFoundError(f"Missing source workbook: {SOURCE_XLSX}")
    if not NOTES_PATH.exists():
        raise FileNotFoundError(f"Missing notes file: {NOTES_PATH}")

    notes_by_file = parse_notes_file(NOTES_PATH)
    json_file_names = relevant_json_file_names()

    workbook = load_workbook(SOURCE_XLSX)
    if "metadata" not in workbook.sheetnames:
        raise ValueError(f"The source workbook has no 'metadata' sheet: {SOURCE_XLSX}")
    worksheet = workbook["metadata"]

    source_columns = [
        worksheet.cell(row=1, column=column_number).value
        for column_number in range(1, len(V4_COLUMNS) + 1)
    ]
    if source_columns != V4_COLUMNS:
        raise ValueError(
            f"Unexpected v4 columns. Expected {V4_COLUMNS}, found {source_columns}."
        )

    workbook_file_names = {
        normalize_file_name(worksheet.cell(row=row_number, column=1).value)
        for row_number in range(2, worksheet.max_row + 1)
    }
    if workbook_file_names != json_file_names:
        missing_from_workbook = sorted(json_file_names - workbook_file_names)
        missing_json_files = sorted(workbook_file_names - json_file_names)
        raise ValueError(
            "The v4 workbook and relevant JSON files do not match. "
            f"Missing from workbook: {missing_from_workbook}; "
            f"missing JSON files: {missing_json_files}."
        )

    notes_column = len(OUTPUT_COLUMNS)
    notes_header = worksheet.cell(row=1, column=notes_column, value="notes")
    source_header = worksheet.cell(row=1, column=notes_column - 1)
    notes_header._style = copy(source_header._style)
    notes_header.alignment = copy(source_header.alignment)

    missing_notes: list[str] = []
    for row_number in range(2, worksheet.max_row + 1):
        file_name = normalize_file_name(worksheet.cell(row=row_number, column=1).value)
        note_body = notes_by_file.get(file_name)
        if note_body is None:
            missing_notes.append(file_name)
            note_text = file_name
        else:
            note_text = f"{file_name}\n{note_body}" if note_body else file_name

        notes_cell = worksheet.cell(row=row_number, column=notes_column, value=note_text)
        notes_cell.alignment = Alignment(vertical="top", wrap_text=True)
        worksheet.row_dimensions[row_number].height = max(
            worksheet.row_dimensions[row_number].height or 15,
            estimate_row_height(note_text),
        )

    worksheet.column_dimensions["L"].width = 100
    worksheet.auto_filter.ref = f"A1:L{worksheet.max_row}"
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT_PATH)

    print(f"Saved {worksheet.max_row - 1} rows to {OUTPUT_PATH}")
    print("Columns:", OUTPUT_COLUMNS)
    print("Missing notes:", len(missing_notes))
    if missing_notes:
        print("Files without notes:")
        for file_name in missing_notes:
            print(f"- {file_name}")


if __name__ == "__main__":
    main()
