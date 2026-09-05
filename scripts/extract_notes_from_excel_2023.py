"""Extract basename, tag, mark, and notes from the 2023 selection workbook."""

from __future__ import annotations

import argparse
import math
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


PROJECT_DIR = Path(
    "/Users/kesju/DI/CODEX_PROJECTS/DUOMENYS_APIE_ATRINKTUS_IRASUS"
)
DEFAULT_INPUT_PATH = (
    PROJECT_DIR / "visi_zive_irasai_annot-Darb_26_06_08_atrankai.xlsx"
)
DEFAULT_OUTPUT_PATH = PROJECT_DIR / "results" / "selected_records_notes_2023.txt"
DEFAULT_SHEET_NAME = "Records"
REQUIRED_COLUMNS = ("basename", "tag", "mark", "notes")
BASENAME_RE = re.compile(r"^\d{7}\.\d{3}$")
WHITESPACE_RE = re.compile(r"\s+")


def normalize_text(value: object) -> str:
    """Convert a cell value to single-line text with compact whitespace."""
    if value is None:
        return ""
    return WHITESPACE_RE.sub(" ", str(value)).strip()


def normalize_basename(value: object) -> str:
    """Normalize a basename to the required XXXXXXX.XXX representation."""
    if isinstance(value, bool):
        basename = str(value)
    elif isinstance(value, (int, float)):
        numeric_value = float(value)
        if not math.isfinite(numeric_value):
            basename = str(value)
        else:
            basename = f"{numeric_value:.3f}"
    else:
        basename = normalize_text(value)

    if basename.casefold().endswith(".json"):
        basename = basename[:-5].strip()
    return basename


def cell_is_struck(cell: Any) -> bool:
    """Return True when a cell or any rich-text run uses strikethrough."""
    if cell.font.strike is True:
        return True

    value = cell.value
    if isinstance(value, str) or value is None:
        return False

    try:
        parts = iter(value)
    except TypeError:
        return False

    for part in parts:
        run_font = getattr(part, "font", None)
        if run_font is not None and getattr(run_font, "strike", None) is True:
            return True
    return False


def find_columns(worksheet: Worksheet) -> tuple[int, dict[str, int]]:
    """Find one header row containing all required column names."""
    for row_number in range(1, min(worksheet.max_row, 25) + 1):
        columns: dict[str, int] = {}
        for cell in worksheet[row_number]:
            header = normalize_text(cell.value).casefold()
            if header in REQUIRED_COLUMNS and header not in columns:
                columns[header] = cell.column

        if all(column in columns for column in REQUIRED_COLUMNS):
            return row_number, columns

    raise ValueError(
        f"Could not find a header row with columns {list(REQUIRED_COLUMNS)} "
        f"in sheet {worksheet.title!r}."
    )


def extract_lines(worksheet: Worksheet) -> tuple[list[str], int, int]:
    """Extract output lines in workbook order and return processing counts."""
    header_row, columns = find_columns(worksheet)
    output_lines: list[str] = []
    skipped_struck = 0
    empty_notes = 0

    for row_number in range(header_row + 1, worksheet.max_row + 1):
        basename_cell = worksheet.cell(row=row_number, column=columns["basename"])
        if basename_cell.value is None or not normalize_text(basename_cell.value):
            continue
        if cell_is_struck(basename_cell):
            skipped_struck += 1
            continue

        basename = normalize_basename(basename_cell.value)
        if not BASENAME_RE.fullmatch(basename):
            raise ValueError(
                f"Unexpected basename {basename!r} in "
                f"{worksheet.title}!{basename_cell.coordinate}."
            )

        tag = normalize_text(
            worksheet.cell(row=row_number, column=columns["tag"]).value
        )
        mark = normalize_text(
            worksheet.cell(row=row_number, column=columns["mark"]).value
        )
        notes = normalize_text(
            worksheet.cell(row=row_number, column=columns["notes"]).value
        )
        if not notes:
            empty_notes += 1

        output_lines.append(" ".join(part for part in (basename, tag, mark, notes) if part))

    return output_lines, skipped_struck, empty_notes


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Extract non-strikethrough basenames with tag, mark, and notes "
            "from the Excel Records sheet."
        )
    )
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT_PATH)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT_PATH)
    parser.add_argument("--sheet", default=DEFAULT_SHEET_NAME)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    if not args.input.exists():
        raise FileNotFoundError(f"Missing source workbook: {args.input}")

    workbook = load_workbook(args.input, data_only=True, rich_text=True)
    if args.sheet not in workbook.sheetnames:
        raise ValueError(
            f"Workbook has no {args.sheet!r} sheet. Available: {workbook.sheetnames}"
        )

    output_lines, skipped_struck, empty_notes = extract_lines(workbook[args.sheet])
    args.output.parent.mkdir(parents=True, exist_ok=True)
    output_text = "\n".join(output_lines)
    if output_text:
        output_text += "\n"
    args.output.write_text(output_text, encoding="utf-8")

    print(f"Saved {len(output_lines)} records to {args.output}")
    print(f"Skipped strikethrough basenames: {skipped_struck}")
    print(f"Records with empty notes: {empty_notes}")


if __name__ == "__main__":
    main()
