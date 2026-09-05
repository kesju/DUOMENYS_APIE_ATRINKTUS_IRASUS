"""Insert 2023 notes and tags into the updated selected-record workbook."""

from __future__ import annotations

import argparse
import math
import re
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


PROJECT_DIR = Path(
    "/Users/kesju/DI/CODEX_PROJECTS/DUOMENYS_APIE_ATRINKTUS_IRASUS"
)
DATA_DIR = Path("/Users/kesju/DI/DUOMENYS/ATRINKTI_ANOTUOTI_IRASAI")
DEFAULT_SOURCE_XLSX = (
    PROJECT_DIR / "results" / "selected_records_metadata_updated_1.xlsx"
)
DEFAULT_NOTES_PATH = PROJECT_DIR / "results" / "selected_records_notes_2023.txt"
DEFAULT_OUTPUT_PATH = (
    PROJECT_DIR / "results" / "selected_records_metadata_updated_with_notes.xlsx"
)
SHEET_NAME = "metadata"
EXPECTED_COLUMNS = [
    "basename",
    "recordingId",
    "userId",
    "comment",
    "tag",
    "rhythm",
    "h_nz_frac%",
    "N",
    "hS",
    "hV",
    "hU",
    "mlS",
    "mlV",
    "mlU",
    "FULLY_ANNOTATED_PROFESSIONALLY",
    "BAD_QUALITY",
    "NOISES_ANNOTATED",
    "Other Flags",
    "notes",
]
BASENAME_AT_LINE_START_RE = re.compile(
    r"^\s*(?P<basename>\d{7}\.\d{3})(?:\.json)?(?:\s+(?P<text>.*?))?\s*$"
)
TAG_RE = re.compile(
    r"(?<!\d)(11110|22220|33330|55550|99990|1111|2222|3333|5555|9999)(?!\d)"
)


def normalize_basename(value: object) -> str:
    """Normalize an Excel or JSON name to the XXXXXXX.XXX identifier."""
    basename = str(value).strip()
    if basename.casefold().endswith(".json"):
        basename = basename[:-5]
    return basename


def is_manifest(file_path: Path) -> bool:
    """Return True for filenames matching manifest*.* case-insensitively."""
    return (
        file_path.name.casefold().startswith("manifest")
        and "." in file_path.name
    )


def relevant_json_basenames() -> set[str]:
    """Return basenames of all JSON records except manifest*.* files."""
    return {
        file_path.stem
        for file_path in DATA_DIR.rglob("*")
        if file_path.is_file()
        and file_path.suffix.casefold() == ".json"
        and not is_manifest(file_path)
    }


def parse_notes_file(path: Path) -> dict[str, str]:
    """Map each line's basename to all single-line text following it."""
    notes_by_basename: dict[str, str] = {}
    for line_number, source_line in enumerate(
        path.read_text(encoding="utf-8").splitlines(), start=1
    ):
        if not source_line.strip():
            continue

        match = BASENAME_AT_LINE_START_RE.fullmatch(source_line)
        if match is None:
            raise ValueError(
                f"Invalid notes line {line_number} in {path}: {source_line!r}"
            )

        basename = match.group("basename")
        if basename in notes_by_basename:
            raise ValueError(
                f"Duplicate basename {basename!r} on line {line_number} in {path}."
            )
        notes_by_basename[basename] = " ".join((match.group("text") or "").split())

    if not notes_by_basename:
        raise ValueError(f"No basename entries found in {path}.")
    return notes_by_basename


def extract_tag_and_notes(text: str) -> tuple[str | None, str]:
    """Extract one supported tag and remove only that occurrence from notes."""
    match = TAG_RE.search(text)
    if match is None:
        return None, text

    tag = match.group(1)
    notes = f"{text[:match.start()]} {text[match.end():]}"
    return tag, " ".join(notes.split())


def validate_worksheet(worksheet: Worksheet) -> tuple[int, int]:
    """Validate the exact schema and return the tag and notes columns."""
    actual_columns = [
        worksheet.cell(row=1, column=column_number).value
        for column_number in range(1, worksheet.max_column + 1)
    ]
    if actual_columns != EXPECTED_COLUMNS:
        raise ValueError(
            "Unexpected source-workbook columns. "
            f"Expected {EXPECTED_COLUMNS}, found {actual_columns}."
        )

    return EXPECTED_COLUMNS.index("tag") + 1, EXPECTED_COLUMNS.index("notes") + 1


def workbook_rows_by_basename(worksheet: Worksheet) -> dict[str, int]:
    """Map unique workbook basenames to row numbers."""
    rows_by_basename: dict[str, int] = {}
    for row_number in range(2, worksheet.max_row + 1):
        basename = normalize_basename(worksheet.cell(row=row_number, column=1).value)
        if not basename:
            raise ValueError(f"Missing basename in Excel row {row_number}.")
        if basename in rows_by_basename:
            raise ValueError(f"Duplicate basename in workbook: {basename!r}.")
        rows_by_basename[basename] = row_number
    return rows_by_basename


def estimate_row_height(text: str, characters_per_line: int = 95) -> float:
    """Estimate a readable height for a wrapped notes cell."""
    line_count = max(1, math.ceil(len(text) / characters_per_line))
    return min(150, 15 * line_count)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Insert matched 2023 notes and extracted tags into Excel."
    )
    parser.add_argument("--input", type=Path, default=DEFAULT_SOURCE_XLSX)
    parser.add_argument("--notes", type=Path, default=DEFAULT_NOTES_PATH)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT_PATH)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    if not args.input.exists():
        raise FileNotFoundError(f"Missing source workbook: {args.input}")
    if not args.notes.exists():
        raise FileNotFoundError(f"Missing notes file: {args.notes}")
    if not DATA_DIR.is_dir():
        raise FileNotFoundError(f"Missing JSON input directory: {DATA_DIR}")

    notes_by_basename = parse_notes_file(args.notes)
    json_basenames = relevant_json_basenames()

    workbook = load_workbook(args.input)
    if SHEET_NAME not in workbook.sheetnames:
        raise ValueError(f"Source workbook has no {SHEET_NAME!r} sheet: {args.input}")
    worksheet = workbook[SHEET_NAME]
    tag_column, notes_column = validate_worksheet(worksheet)
    rows_by_basename = workbook_rows_by_basename(worksheet)

    workbook_basenames = set(rows_by_basename)
    if workbook_basenames != json_basenames:
        raise ValueError(
            "The source workbook and relevant JSON files do not match. "
            f"Missing from workbook: {sorted(json_basenames - workbook_basenames)}; "
            f"missing JSON files: {sorted(workbook_basenames - json_basenames)}."
        )

    matched_basenames = workbook_basenames & set(notes_by_basename)
    tags_found = 0
    for basename, row_number in rows_by_basename.items():
        if basename not in matched_basenames:
            continue

        tag, notes = extract_tag_and_notes(notes_by_basename[basename])
        worksheet.cell(row=row_number, column=tag_column).value = tag
        worksheet.cell(row=row_number, column=notes_column).value = notes or None
        worksheet.row_dimensions[row_number].height = max(
            worksheet.row_dimensions[row_number].height or 15,
            estimate_row_height(notes),
        )
        if tag is not None:
            tags_found += 1

    args.output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(args.output)

    print(f"Saved {len(rows_by_basename)} rows to {args.output}")
    print(f"Rows matched to notes: {len(matched_basenames)}")
    print(f"Tags extracted: {tags_found}")
    print(f"Workbook rows left unchanged: {len(rows_by_basename) - len(matched_basenames)}")
    print(f"Unused notes entries: {len(set(notes_by_basename) - workbook_basenames)}")


if __name__ == "__main__":
    main()
